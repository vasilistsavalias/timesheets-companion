import io
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

class GreekExcelExporter:
    def __init__(self, data):
        self.data = data

    def export_excel_stream(self):
        """Generates a professional timesheet log in memory."""
        output = io.BytesIO()
        df = pd.DataFrame(self.data)
        
        # Use pandas with openpyxl engine
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Detailed Log')
            ws = writer.sheets['Detailed Log']
            
            # Formatting
            for col_idx, column in enumerate(df.columns, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = Font(bold=True, name='Arial')
                cell.fill = PatternFill(start_color='D3D3D3', fill_type='solid')
                
                # Auto-width
                max_length = max(df[column].astype(str).map(len).max(), len(column)) + 4
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length, 80)

        return output.getvalue()

    def update_niki_excel_stream(self, niki_file_stream, month_name, ee_totals, detailed_schedule):
        """
        Updates Niki's original Excel in-memory.
        niki_file_stream: a file-like object from st.file_uploader
        """
        output = io.BytesIO()
        # Load from stream
        wb = load_workbook(niki_file_stream)
        ws = wb.active
        
        # 1. Update Summary Column
        month_col = None
        for cell in ws[1]:
            if cell.value and str(cell.value).strip().upper() == month_name.upper():
                month_col = cell.column
                break
        
        if month_col:
            for ee_label, total_hours in ee_totals.items():
                for row in range(1, 25):
                    cell_val = str(ws.cell(row=row, column=1).value).upper()
                    if ee_label.upper() in cell_val:
                        ws.cell(row=row + 2, column=month_col, value=total_hours)
                        ws.cell(row=row + 2, column=month_col).font = Font(bold=True, color="0000FF")
                        break

        # 2. Append Detailed Log starting at Row 30
        start_row = 30
        ws.cell(row=start_row, column=1, value=f"DETAILED LOG FOR {month_name.upper()}").font = Font(bold=True, size=12)
        
        headers = ["Date", "Deliverable", "Hours", "Description"]
        for i, h in enumerate(headers, start=1):
            cell = ws.cell(row=start_row + 1, column=i, value=h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='E0E0E0', fill_type='solid')

        for i, entry in enumerate(detailed_schedule, start=1):
            ws.cell(row=start_row + 1 + i, column=1, value=entry['Date'])
            ws.cell(row=start_row + 1 + i, column=2, value=entry['Deliverable'])
            ws.cell(row=start_row + 1 + i, column=3, value=entry['Hours'])
            ws.cell(row=start_row + 1 + i, column=4, value=entry['Description'])
            
        wb.save(output)
        return output.getvalue()